from openpyxl import load_workbook

wb=load_workbook('lessons.xlsx')

t=wb.get_sheet_by_name("Worksheet")

print(t['A1'].value)
print(t['A1'].column)
marks=[]

def analyze(s):
    number=0
    value=0
    for c in s:
        if c in ['1','2','3','4','5']:
            print(c)
            value=value+int(c)
            number+=1
    return(number,value)

for row in range(4,24):
    r=str(row)
    


    subject=t['A'+r].value
    
    print(subject)
    number=0
    value=0

    for col in range(2,123):
        c=str(t.cell(row=row,column=col).value)
        if(c):

            analyz=analyze(c)
            number+=analyz[0]
            value+=analyz[1]
    

    if(number!=0):
        mean=value/number
    else:
        mean=0
    marks.append({subject:mean})

    


    


for i in marks:
    print(i)
        

    